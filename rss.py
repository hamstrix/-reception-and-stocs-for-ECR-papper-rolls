#RECEPTIE SI EVIDENTA STOCURI ROLE
from openpyxl import Workbook
from openpyxl import styles
def banner(message, border = "*"):
    line = border * len(message)
    print(line)
    print(message)
    print(line)

#Introducem optiunile de Receptie si Evidenta stoc 
print("Alegeti operatiunea dorita :\n\n1 Receptie Role\n2 Evidenta stoc role\n\n")
operator = input("operatiunea:\n")

#intram in modulul Receptie Role
while operator =="1" :

    banner ("RECEPTIE ROLE" ,"*")
    
#introducem cele doua firme pe care se face receptia rolelor
    print ("\n\nSelecteaza firma dorita:\n\n1 DANUBIUS EXIM\n2 DATA LOGIC SYSTEMS\n\n")
    firma = input ("optiunea ta:\n")
    
# intram in modulul receptie role sectiunea Danubius Exim
    
    if firma == "1":
        doc = open (r"C:\Users\Tehnic 7\Desktop\program scanare\provizoriu_receptie.txt" ,"w")
        doc.writelines("TIP ROLA;U.M.;COD PRODUS;COLETARE CUTIE;NR.CUTII FIZIC;NR. BUCATI FIZIC;TOTAL BUCATI\n")
        while True :
            
            print("Alegeti tipul de role dorit:\n\n1 Role 35mm\20mmx18m(70088)\n2 Role 35mm\20mmx30m(70087)\n3 Role 27.5mm\27.5mmx18m(70085)\n4 Role 27.5mm\27.5mmx30m(70089)\n5 Role 37.5mmx30m(70095)\n6 Role 57mmx18m(70091)\n7 Role 57mmx20m(70093)\n8 Role 57mmx40m(70086)\n9 Role 79mmx34m(72169)\n10 Role 79mmx40m(72163)\n11 Role 80mmx34m(71724)\n12 Role 80mmx40m(70092)\n13 Role 57mm x 30mDatalogic(70988)\n14 role 56mm x 20m (70084)\n\n")
            rola = input ("Alegeti tipul de rola:")
            
            if rola =="1" :
                print ( "Role 35mmx20mmx18m(70088)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 120*cutii+bucati
                p = "35mmx20mmx18m;set;70088;120;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
            
            elif rola =="2" :
                print ( "Role 35mmx20mmx30m(70087)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 80*cutii+bucati
                p = "35mmx20mmx30m;set;70087;80;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
                
            elif rola =="3" :
                print ( "Role 27.5mmx27.5mmx18m(70085)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 120*cutii+bucati
                p = "27.5mmx27.5mmx18m;set;70085;240;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
            
            elif rola =="4" :
                print ( "Role 27.5mmx27.5mmx30m(70089)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 160*cutii+bucati
                p = "27.5mmx27.5mmx30m;set;70089;160;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
                
            elif rola =="5" :
                print ( "Role 37.5mmx37.5mmx30m(70095)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 120*cutii+bucati
                p = "37.5mmx37.5mmx30m;set;70095;120;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
            
            elif rola =="6" :
                print ( "Role 57mmx18m(70091)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 160*cutii+bucati
                p = "57mmx18m;buc;70091;160;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p) 
            
            elif rola =="7" :
                print ( "Role 57mmx20m(70093)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 120*cutii+bucati
                p = "57mmx20m;buc;70093;120;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)    
            
            elif rola =="8" :
                print ( "Role 57mmx40m(70086)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 80*cutii+bucati
                p = "57mmx40m;buc;70086;80;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
                
            elif rola =="9" :
                print ( "Role 79mmx34m(72169)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 60*cutii+bucati
                p = "79mmx34m;buc;2169;60;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
                
            elif rola =="10" :
                print ( "Role 79mmx40m(72163)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 60*cutii+bucati
                p = "79mmx40m;set;72163;60;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
                
            elif rola =="11" :
                print ( "Role 80mmx34m(71724)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 60*cutii+bucati
                p = "80mmx34m;set;71724;60;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
            
            elif rola =="12" :
                print ( "Role 80mmx40m(70092)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 60*cutii+bucati
                p = "80mmx40m;set;70092;60;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
                
            elif rola =="13" :
                print ( "Role 57mmx30m Datalogic(70988)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 80*cutii+bucati
                p = "57mmx30m;set;70988;80;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
            
            elif rola =="14" :
                print ( "Role 56mmx20m(70084)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 72*cutii+bucati
                p = "56mmx20m;buc;70084;72;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
            
           
            else :
                doc.close()
                document  = open("C:\\users\\tehnic 7\\Desktop\\program scanare\\provizoriu_receptie.txt")
                #deschidem o lista in care vor fi introduse toate intrarile din fisier
                intrari=[]
                #ne asiguram ca parcurgem documentul de la inceput
                document.seek(0)
                #impartim fiecare rand prin " " 

                for intrare in document.readlines():
                    intrari.append(intrare.rstrip("\n").split(";"))
                #printam lista intrarilor
                print(intrari ,"\n")
                #Deschidem un workbook nou

                tabel = Workbook()
                path = "C:\\Users\\Tehnic 7\\Desktop\\receptie marfa\\rollco\\"+"DANUBIUS "+input("Introduceti data receptiei urmat de .xlsx :\n")
                tabel.save(path)
                w = tabel.create_sheet("Receptie Rollco  - Danubius" ,0 )
                for row in intrari:
                    w = tabel["Receptie Rollco  - Danubius"]
                    w.append(row)
                    w.auto_filter.ref = "A1:G2000"
                    w.auto_filter.add_filter_column(0, ["35mmx20mmx18m(70088)"])
                    w.auto_filter.add_sort_condition("A2:A2000")
                    w.column_dimensions["A"].width =20
                    w.column_dimensions["B"].width =18                    
                    w.column_dimensions["C"].width =18                    
                    w.column_dimensions["D"].width =18                    
                    w.column_dimensions["E"].width =18                    
                    w.column_dimensions["F"].width =18                    
                    w.column_dimensions["G"].width =18
                    
                tabel.save(path)
                tabel.close()
                document.close()
                break
        
################################################

    if firma == "2":
        doc = open (r"C:\Users\Tehnic 7\Desktop\program scanare\provizoriu_receptie.txt" ,"w")
        doc.writelines("TIP ROLA;U.M.;COD PRODUS;COLETARE CUTIE;NR.CUTII FIZIC;NR. BUCATI FIZIC;TOTAL BUCATI\n")
        while True :
            
            print("Alegeti tipul de role dorit:\n\n1 Role 35mm\20mmx18m(70088)\n2 Role 35mm\20mmx30m(70087)\n3 Role 27.5mm\27.5mmx18m(70085)\n4 Role 27.5mm\27.5mmx30m(70089)\n5 Role 37.5mmx30m(70095)\n6 Role 57mmx18m(70091)\n7 Role 57mmx20m(70093)\n8 Role 57mmx40m(70086)\n9 Role 79mmx34m(72169)\n10 Role 79mmx40m(72163)\n11 Role 80mmx34m(71724)\n12 Role 80mmx40m(70092)\n13 Role 57mm x 30mDatalogic(70988)\n14 role 56mm x 20m (70084)\n\n")
            rola = input ("Alegeti tipul de rola:")
            
            if rola =="1" :
                print ( "Role 35mmx20mmx18m(70088)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 120*cutii+bucati
                p = "35mmx20mmx18m;set;70088;120;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
            
            elif rola =="2" :
                print ( "Role 35mmx20mmx30m(70087)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 80*cutii+bucati
                p = "35mmx20mmx30m;set;70087;80;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
                
            elif rola =="3" :
                print ( "Role 27.5mmx27.5mmx18m(70085)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 120*cutii+bucati
                p = "27.5mmx27.5mmx18m;set;70085;240;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
            
            elif rola =="4" :
                print ( "Role 27.5mmx27.5mmx30m(70089)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 160*cutii+bucati
                p = "27.5mmx27.5mmx30m;set;70089;160;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
                
            elif rola =="5" :
                print ( "Role 37.5mmx37.5mmx30m(70095)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 120*cutii+bucati
                p = "37.5mmx37.5mmx30m;set;70095;120;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
            
            elif rola =="6" :
                print ( "Role 57mmx18m(70091)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 120*cutii+bucati
                p = "57mmx18m;buc;70091;120;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p) 
            
            elif rola =="7" :
                print ( "Role 57mmx20m(70093)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 120*cutii+bucati
                p = "57mmx20m;buc;70093;120;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)    
            
            elif rola =="8" :
                print ( "Role 57mmx40m(70086)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 80*cutii+bucati
                p = "57mmx40m;buc;70086;80;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
                
            elif rola =="9" :
                print ( "Role 79mmx34m(72169)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 60*cutii+bucati
                p = "79mmx34m;buc;72169;60;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
                
            elif rola =="10" :
                print ( "Role 79mmx40m(72163)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 60*cutii+bucati
                p = "79mmx40m;buc;72163;60;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
                
            elif rola =="11" :
                print ( "Role 80mmx34m(71724)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 60*cutii+bucati
                p = "80mmx34m;buc;71724;60;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
            
            elif rola =="12" :
                print ( "Role 80mmx40m(70092)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 60*cutii+bucati
                p = "80mmx40m;buc;70092;60;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
                
            elif rola =="13" :
                print ( "Role 57mmx30m Datalogic(70988)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 80*cutii+bucati
                p = "57mmx30m;buc;70988;80;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
            
            elif rola =="14" :
                print ( "Role 56mmx20m(70084)\n")
                cutii = int(input("Introduceti numarul fizic de cutii:"))
                bucati = int(input ("Introduceti numarul fizic de bucati/seturi:"))
                final  = 72*cutii+bucati
                p = "56mmx20m;buc;70084;72;"+str(cutii)+";"+str(bucati)+";"+str(final)+"\n"
                doc.writelines(p)
            
            
            else :
               
                doc.close()
                document  = open("C:\\users\\tehnic 7\\Desktop\\program scanare\\provizoriu_receptie.txt")
                #deschidem o lista in care vor fi introduse toate intrarile din fisier
                intrari=[]
                #ne asiguram ca parcurgem documentul de la inceput
                document.seek(0)
                #impartim fiecare rand prin " " 

                for intrare in document.readlines():
                    intrari.append(intrare.rstrip("\n").split(";"))
                #printam lista intrarilor
                print(intrari ,"\n")
                #Deschidem un workbook nou

                tabel = Workbook()
                path = "C:\\Users\\Tehnic 7\\Desktop\\receptie marfa\\rollco\\"+"DATALOGIC "+input("Introduceti data receptiei urmat de .xlsx :\n")
                tabel.save(path)
                w = tabel.create_sheet("Receptie Rollco  - Datalogic",0 )
                for row in intrari:
                    w = tabel["Receptie Rollco  - Datalogic"]
                    w.append(row)
                    w.auto_filter.ref = "A1:G2000"
                    w.auto_filter.add_filter_column(0, ["35mmx20mmx18m(70088)"])
                    w.auto_filter.add_sort_condition("A2:A2000")
                    w.column_dimensions["A"].width =20
                    w.column_dimensions["B"].width =18
                    w.column_dimensions["C"].width =18
                    w.column_dimensions["D"].width =18
                    w.column_dimensions["E"].width =18
                    w.column_dimensions["F"].width =18
                    w.column_dimensions["G"].width =18
                    
                tabel.save(path)
                tabel.close()
                document.close()
                break
    inapoi = input ("Selectati si alt distribuitor pentru receptie? (y/n)")
    if inapoi == "y" :
        continue
    else :
        break

# Intram in modulul de Stocuri Rollco

if operator =="2":
    
#Introducem bannerul de recunoastere
    banner("STOCURI DANUBIUS-ROLLCO" , "*")
 #Deschidem un document provizoriu in care trecem toate valorile stocului   
    doc = open (r"C:\Users\Tehnic 7\Desktop\program scanare\provizoriu_stoc.txt" ,"w")
    doc.writelines("TIP ROLA;U.M.;COD PRODUS;COLETARE CUTIE;NR.CUTII FIZIC;NR. BUCATI FIZIC;TOTAL BUCATI FIZIC;TOTAL BUCATI SCRIPTIC;DIFERENTE\n")
 
 #Intram in fiecare tip de rola in parte
 
    print ("\n\n\nRole 56mmx20m (cod: 70084)\n\n")
    scriptic = input ("Cantitatea scriptica de role seturi:")
    cutii = input ("Cantitatea fizica de cutii :")
    bucati = input ("Cantitatea fizica de set-uri:")
    total =72*int(cutii)+int(bucati)
    diferenta = int(total)-int(scriptic)
    p= "56mmx20m;buc;70084;72;"+str(cutii)+";"+str(bucati)+";"+str(total)+";"+str(scriptic)+";"+str(diferenta)+"\n"
    doc.writelines(p)
    
    print ("\n\n\nRole 35mmx20mmx18m (cod: 70088)\n\n")
    scriptic = input ("Cantitatea scriptica de role seturi:")
    cutii = input ("Cantitatea fizica de cutii :")
    bucati = input ("Cantitatea fizica de set-uri:")
    total =120*int(cutii)+int(bucati)
    diferenta = int(total)-int(scriptic)
    p= "35mmx20mmx18m;set;70088;120;"+str(cutii)+";"+str(bucati)+";"+str(total)+";"+str(scriptic)+";"+str(diferenta)+"\n"
    doc.writelines(p)
    
    print ("\n\n\nRole 35mmx20mmx30m (cod: 70087)\n\n")
    scriptic = input ("Cantitatea scriptica de role seturi:")
    cutii = input ("Cantitatea fizica de cutii :")
    bucati = input ("Cantitatea fizica de set-uri:")
    total =80*int(cutii)+int(bucati)
    diferenta = int(total)-int(scriptic)
    p= "35mmx20mmx30m;set;70087;80;"+str(cutii)+";"+str(bucati)+";"+str(total)+";"+str(scriptic)+";"+str(diferenta)+"\n"
    doc.writelines(p)
    
    print ("\n\n\nRole 27.5mmx27.5mmx18m (cod: 70085)\n\n")
    scriptic = input ("Cantitatea scriptica de role seturi:")
    cutii = input ("Cantitatea fizica de cutii :")
    bucati = input ("Cantitatea fizica de set-uri:")
    total =120*int(cutii)+int(bucati)
    diferenta = int(total)-int(scriptic)
    p= "27.5mmx27.5mmx18m;set;70085;240;"+str(cutii)+";"+str(bucati)+";"+str(total)+";"+str(scriptic)+";"+str(diferenta)+"\n"
    doc.writelines(p)
    
    print ("\n\n\nRole 27.5mmx27.5mmx30m (cod: 70089)\n\n")
    scriptic = input ("Cantitatea scriptica de role seturi:")
    cutii = input ("Cantitatea fizica de cutii :")
    bucati = input ("Cantitatea fizica de set-uri:")
    total =160*int(cutii)/2+int(bucati)
    diferenta = int(total)-int(scriptic)
    p= "27.5mmx27.5mmx30m;set;70089;160;"+str(cutii)+";"+str(bucati)+";"+str(total)+";"+str(scriptic)+";"+str(diferenta)+"\n"
    doc.writelines(p)
    
    print ("\n\n\nRole 37.5mmx37.5mmx30m (cod: 70095)\n\n")
    scriptic = input ("Cantitatea scriptica de role seturi:")
    cutii = input ("Cantitatea fizica de cutii :")
    bucati = input ("Cantitatea fizica de set-uri:")
    total =120*int(cutii)/2+int(bucati)
    diferenta = int(total)-int(scriptic)
    p= "37.5mmx37.5mmx30m;set;70095;120;"+str(cutii)+";"+str(bucati)+";"+str(total)+";"+str(scriptic)+";"+str(diferenta)+"\n"
    doc.writelines(p)
    
    print ("\n\n\nRole 57mmx18m (cod: 70091)\n\n")
    scriptic = input ("Cantitatea scriptica de role seturi:")
    cutii = input ("Cantitatea fizica de cutii :")
    bucati = input ("Cantitatea fizica de set-uri:")
    total =160*int(cutii)+int(bucati)
    diferenta = int(total)-int(scriptic)
    p= "57mmx18m;buc;70091;160;"+str(cutii)+";"+str(bucati)+";"+str(total)+";"+str(scriptic)+";"+str(diferenta)+"\n"
    doc.writelines(p)
    
    print ("\n\n\nRole 57mmx20m (cod: 70093)\n\n")
    scriptic = input ("Cantitatea scriptica de role seturi:")
    cutii = input ("Cantitatea fizica de cutii :")
    bucati = input ("Cantitatea fizica de set-uri:")
    total =120*int(cutii)+int(bucati)
    diferenta = int(total)-int(scriptic)
    p= "57mmx20m;buc;70093;120;"+str(cutii)+";"+str(bucati)+";"+str(total)+";"+str(scriptic)+";"+str(diferenta)+"\n"
    doc.writelines(p)
    
    print ("\n\n\nRole 57mmx40m (cod: 70086)\n\n")
    scriptic = input ("Cantitatea scriptica de role seturi:")
    cutii = input ("Cantitatea fizica de cutii :")
    bucati = input ("Cantitatea fizica de set-uri:")
    total =80*int(cutii)+int(bucati)
    diferenta = int(total)-int(scriptic)
    p= "57mmx40m;buc;70086;800;"+str(cutii)+";"+str(bucati)+";"+str(total)+";"+str(scriptic)+";"+str(diferenta)+"\n"
    doc.writelines(p)
    
    print ("\n\n\nRole 79mmx34m (cod: 72169)\n\n")
    scriptic = input ("Cantitatea scriptica de role seturi:")
    cutii = input ("Cantitatea fizica de cutii :")
    bucati = input ("Cantitatea fizica de set-uri:")
    total =60*int(cutii)+int(bucati)
    diferenta = int(total)-int(scriptic)
    p= "79mmx34m;buc;72169;60;"+str(cutii)+";"+str(bucati)+";"+str(total)+";"+str(scriptic)+";"+str(diferenta)+"\n"
    doc.writelines(p)
    
    print ("\n\n\nRole 79mmx40m (cod: 72163)\n\n")
    scriptic = input ("Cantitatea scriptica de role seturi:")
    cutii = input ("Cantitatea fizica de cutii :")
    bucati = input ("Cantitatea fizica de set-uri:")
    total =60*int(cutii)+int(bucati)
    diferenta = int(total)-int(scriptic)
    p= "79mmx40m;buc;72163;60;"+str(cutii)+";"+str(bucati)+";"+str(total)+";"+str(scriptic)+";"+str(diferenta)+"\n"
    doc.writelines(p)
    
    print ("\n\n\nRole 80mmx34m (cod: 71724)\n\n")
    scriptic = input ("Cantitatea scriptica de role seturi:")
    cutii = input ("Cantitatea fizica de cutii :")
    bucati = input ("Cantitatea fizica de set-uri:")
    total =60*int(cutii)+int(bucati)
    diferenta = int(total)-int(scriptic)
    p= "80mmx34m;buc;71724;60;"+str(cutii)+";"+str(bucati)+";"+str(total)+";"+str(scriptic)+";"+str(diferenta)+"\n"
    doc.writelines(p)
    
    print ("\n\n\nRole 80mmx40m (cod: 70092)\n\n")
    scriptic = input ("Cantitatea scriptica de role seturi:")
    cutii = input ("Cantitatea fizica de cutii :")
    bucati = input ("Cantitatea fizica de set-uri:")
    total =60*int(cutii)+int(bucati)
    diferenta = int(total)-int(scriptic)
    p= "80mmx40m;buc;70092;60;"+str(cutii)+";"+str(bucati)+";"+str(total)+";"+str(scriptic)+";"+str(diferenta)+"\n"
    doc.writelines(p)
    
    print ("\n\n\nRole 57mmx30m DATALOGIC (cod: 70988)\n\n")
    scriptic = input ("Cantitatea scriptica de role seturi:")
    cutii = input ("Cantitatea fizica de cutii :")
    bucati = input ("Cantitatea fizica de set-uri:")
    total =80*int(cutii)+int(bucati)
    diferenta = int(total)-int(scriptic)
    p= "57mmx30m DATALOGIC;buc;70988;80;"+str(cutii)+";"+str(bucati)+";"+str(total)+";"+str(scriptic)+";"+str(diferenta)+"\n"
    doc.writelines(p)
    doc.close()
                
document  = open("C:\\users\\tehnic 7\\Desktop\\program scanare\\provizoriu_stoc.txt")
#deschidem o lista in care vor fi introduse toate intrarile din fisier
intrari=[]
#ne asiguram ca parcurgem documentul de la inceput
document.seek(0)
#impartim fiecare rand prin " " 

for intrare in document.readlines():
    intrari.append(intrare.rstrip("\n").split(";"))
    #printam lista intrarilor
    print(intrari ,"\n")
#Deschidem un workbook nou

tabel = Workbook()
path = "C:\\Users\\Tehnic 7\\Desktop\\stocuri\\stoc role\\"+"DANUBIUS-ROLLCO "+input("Introduceti data la care s-a efectuat stocul urmat de .xlsx :\n")
tabel.save(path)
w = tabel.create_sheet("Rollco",0 )
for row in intrari:
    w = tabel["Rollco"]
    w.append(row)
    w.auto_filter.ref = "A1:I2000"
    w.auto_filter.add_filter_column(0, ["56mmx20m"])
    w.auto_filter.add_sort_condition("A2:A2000")
    w.column_dimensions["A"].width =24
    w.column_dimensions["B"].width =18
    w.column_dimensions["C"].width =18
    w.column_dimensions["D"].width =18
    w.column_dimensions["E"].width =18
    w.column_dimensions["F"].width =18
    w.column_dimensions["G"].width =18
    w.column_dimensions["H"].width =18
    w.column_dimensions["I"].width =18                    
    tabel.save(path)
    
tabel.close()
document.close()